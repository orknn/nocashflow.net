#!/usr/bin/env python3
"""NoCashFlow · macro "the regime" snapshot → data/macro2.json

Free, server-side. Auto-sourced:
  · FRED   — CPI/Core CPI/Core PCE (→ YoY), 5y5y breakeven, Fed target, balance
             sheet, RRP, M2, net liquidity (WALCL − TGA − RRP), jobless claims,
             nonfarm payrolls (→ MoM change), the Treasury curve, 2s10s, HY spread.
  · DefiLlama — stablecoin supply.
  · Atlanta Fed — GDPNow (xlsx; needs openpyxl).

Manual fields (no free API — kept from the last-good file, edited by hand):
  Fed rate-cut odds, dot-plot median, ISM, next-FOMC date, global M2.
Editorial read/notes live in macro-notes.json (separate, human-committed).
Resilience: each series falls back to last-good; nothing is fabricated.
"""
import csv
import io
import json
import re
from collections import defaultdict
from datetime import date, datetime, timezone
from pathlib import Path

import requests

KH = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120 Safari/537.36",
      "Accept": "application/json"}
_MON = {"JAN": 1, "FEB": 2, "MAR": 3, "APR": 4, "MAY": 5, "JUN": 6,
        "JUL": 7, "AUG": 8, "SEP": 9, "OCT": 10, "NOV": 11, "DEC": 12}

DATA = Path(__file__).resolve().parent.parent / "data"
UA = {"User-Agent": "Mozilla/5.0 (NoCashFlow data fetcher)"}
TIMEOUT = 20
FRED = "https://fred.stlouisfed.org/graph/fredgraph.csv"


# FOMC decision dates (2nd day) — schedule is published a year ahead; "which is
# next" is computed by date so it advances on its own as meetings pass.
FOMC_DATES = ["2026-01-28", "2026-03-18", "2026-04-29", "2026-06-17", "2026-07-29",
              "2026-09-16", "2026-10-28", "2026-12-09", "2027-01-27", "2027-03-17"]


def now_iso():
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def _next_fomc_dates(n=4):
    today = date.today()
    out = [date.fromisoformat(d) for d in FOMC_DATES if date.fromisoformat(d) > today]
    return out[:n]


def _next_fomc():
    nd = _next_fomc_dates(1)
    if not nd:
        return "—", ""
    dd = nd[0]
    return f"{dd.strftime('%b')} {dd.day}", f"{(dd - date.today()).days}d"


def fetch_cut_odds():
    """Rate-cut probability for upcoming FOMC meetings from Kalshi (keyless,
    prediction-market implied). Per meeting: sum the 'Cut' bucket yes-prices
    (cents == %). Returns [{m, p}] for the next 4 meetings, or None."""
    try:
        r = requests.get("https://api.elections.kalshi.com/trade-api/v2/markets",
                         params={"series_ticker": "KXFEDDECISION", "status": "open", "limit": 1000},
                         headers=KH, timeout=TIMEOUT)
        r.raise_for_status()
        markets = r.json().get("markets", [])
    except Exception as e:
        print(f"  ⚠️  Kalshi cut-odds: {e}")
        return None
    cut = defaultdict(float)                          # (year, month) → summed cut %
    diag_done = False
    for m in markets:
        parts = m.get("ticker", "").split("-")
        if len(parts) < 3 or not parts[2].startswith("C"):
            continue                                  # cut buckets only (C25, C26…)
        mm = re.match(r"(\d{2})([A-Z]{3})", parts[1])  # ticker date is YYMon (2026=year!)
        if not mm or mm.group(2) not in _MON:
            continue
        if not diag_done:
            print("  [diag2] cut market fields:", {k: m.get(k) for k in
                  ("ticker", "last_price", "yes_bid", "yes_ask", "volume", "status")})
            diag_done = True
        price = m.get("last_price") or m.get("yes_ask") or m.get("yes_bid") or 0
        cut[(2000 + int(mm.group(1)), _MON[mm.group(2)])] += price
    print(f"  [diag] cut keys={dict(cut)} | want={[(d.year, d.month) for d in _next_fomc_dates(4)]}")
    odds = []
    for dd in _next_fomc_dates(4):
        if (dd.year, dd.month) in cut:
            odds.append({"m": f"{dd.strftime('%b')} {dd.day}", "p": round(cut[(dd.year, dd.month)])})
    return odds or None


def load_json(name, default):
    p = DATA / name
    if p.exists():
        try:
            return json.loads(p.read_text(encoding="utf-8"))
        except Exception:
            pass
    return default


def write_json(name, obj):
    (DATA / name).write_text(json.dumps(obj, ensure_ascii=False, indent=2) + "\n",
                             encoding="utf-8")


def fred_series(sid, start="2022-01-01"):
    """[(date, float)] oldest→newest, missing dropped, or [] on failure."""
    try:
        r = requests.get(FRED, params={"id": sid, "cosd": start}, headers=UA, timeout=TIMEOUT)
        r.raise_for_status()
        out = []
        for row in list(csv.reader(io.StringIO(r.text)))[1:]:
            if len(row) >= 2 and row[1] not in (".", ""):
                out.append((row[0], float(row[1])))
        return out
    except Exception as e:
        print(f"  ⚠️  FRED {sid}: {e}")
        return []


def fred_last(sid):
    s = fred_series(sid)
    return s[-1][1] if s else None


def fred_yoy(sid):
    """Year-over-year % from a monthly index series (last vs ~12 months prior)."""
    s = fred_series(sid)
    if len(s) < 13:
        return None
    return round((s[-1][1] / s[-13][1] - 1) * 100, 1)


def _fred_nearest_year(sid):
    """Value for the current calendar year from an annual-projection FRED series
    (e.g. FEDTARMD dot plot: rows dated 2026/2027/2028 → take 2026)."""
    s = fred_series(sid, start="2025-01-01")
    yr = str(datetime.now(timezone.utc).year)
    for d, v in s:
        if d.startswith(yr):
            return v
    return s[-1][1] if s else None


def _arrow(curr, prev):
    if prev is None or curr is None:
        return "", "neu"
    if curr > prev:
        return "▲", "up"
    if curr < prev:
        return "▼", "down"
    return "→", "neu"


def st(v, sub="", dir="neu"):
    return {"v": v, "sub": sub, "dir": dir}


def build_macro2():
    prev = load_json("macro2.json", {})
    out = dict(prev)  # preserve manual fields; overwrite what we source
    out["_note"] = "Live: FRED + DefiLlama + Atlanta Fed. Manual (no free API): " \
                   "cut-odds, dot plot, ISM, next-FOMC, global M2."
    out["updated"] = now_iso()

    # ── inflation (YoY from index series) ────────────────────────────────────
    cpi_h, cpi_c, pce_c = fred_yoy("CPIAUCSL"), fred_yoy("CPILFESL"), fred_yoy("PCEPILFE")
    be = fred_last("T5YIFR")
    pinf = prev.get("inflation", [{}, {}, {}, {}])

    def infl(curr, prev_stat):
        pv = _num(prev_stat.get("v"))
        sub, d = _arrow(curr, pv)
        # green when inflation cools (down), red when it heats up — Fed-friendly
        d = {"up": "down", "down": "up"}.get(d, "neu")
        sub_txt = f"{sub} from {pv:.1f}%" if (pv is not None and sub) else ""
        return st(f"{curr:.1f}%", sub_txt, d)

    if cpi_h is not None:
        pinf[0] = infl(cpi_h, pinf[0] if len(pinf) > 0 else {})
    if cpi_c is not None:
        pinf[1] = infl(cpi_c, pinf[1] if len(pinf) > 1 else {})
    if pce_c is not None:
        pinf[2] = infl(pce_c, pinf[2] if len(pinf) > 2 else {})
    if be is not None:
        pv = _num((pinf[3] if len(pinf) > 3 else {}).get("v"))
        sub, d = _arrow(be, pv)
        pinf[3] = st(f"{be:.2f}%", f"{sub} {'+' if (pv and be>=pv) else ''}{round((be-pv)*100):.0f}bp" if pv else "", d)
    out["inflation"] = pinf

    # ── fed (balance sheet, RRP, target; cut_odds & dots stay manual) ────────
    fed = dict(prev.get("fed", {}))
    fstats = list(fed.get("stats", [st(""), st(""), st(""), st("")]))
    lo, hi = fred_last("DFEDTARL"), fred_last("DFEDTARU")
    if lo is not None and hi is not None:
        fstats[0] = st(f"{lo:.2f}–{hi:.2f}", fstats[0].get("sub", "held"), "neu")
    bs = fred_last("WALCL")
    if bs is not None:
        fstats[1] = st(f"${bs / 1e6:.1f}T", "▼ QT ongoing", "down")
    rrp = fred_last("RRPONTSYD")
    if rrp is not None:
        fstats[2] = st(f"${rrp:.0f}B" if rrp >= 1 else f"${rrp*1000:.0f}M", "▼ draining", "down")
    # dot-plot: nearest-year SEP median fed funds projection (FRED FEDTARMD)
    dot = _fred_nearest_year("FEDTARMD")
    if dot is not None:
        fstats[3] = st(f"{dot:.2f}%", f"{datetime.now(timezone.utc).year} median", "neu")
    fed["stats"] = fstats
    cut_odds = fetch_cut_odds()                       # Kalshi prediction-market implied
    if cut_odds:
        fed["cut_odds"] = cut_odds
    out["fed"] = fed

    # ── liquidity (M2, net liquidity = WALCL − TGA − RRP, + stablecoin) ──────
    liq = dict(prev.get("liquidity", {}))
    liq.pop("global_m2", None)                        # US M2 is enough (per project)
    lstats = [st(""), st(""), st("")]                 # M2 · net liquidity · stablecoins
    m2 = fred_last("M2SL")
    m2yoy = fred_yoy("M2SL")
    if m2 is not None:
        lstats[0] = st(f"${m2 / 1000:.1f}T",
                       (f"▲ +{m2yoy:.1f}% YoY" if (m2yoy or 0) >= 0 else f"▼ {m2yoy:.1f}% YoY") if m2yoy is not None else "",
                       "up" if (m2yoy or 0) >= 0 else "down")
    # net liquidity series (weekly): WALCL − TGA(WTREGEN) − RRP
    wal = dict(fred_series("WALCL"))
    tga = dict(fred_series("WTREGEN"))
    rrps = dict(fred_series("RRPONTSYD"))
    # units: WALCL & WTREGEN are $millions, RRPONTSYD is $billions → normalise to $T
    series = []
    for d in sorted(wal):
        if d in tga and d in rrps:
            series.append(round((wal[d] - tga[d] - rrps[d] * 1000) / 1e6, 2))  # $T
    if len(series) >= 8:
        series = series[-26:]  # ~6 months weekly
        liq["net_liq_series"] = series
        liq["y_min"] = round(min(series) - 0.1, 1)
        liq["y_max"] = round(max(series) + 0.1, 1)
        nl = series[-1]
        nl3 = round((nl / series[-13] - 1) * 100, 1) if len(series) > 13 else None
        lstats[1] = st(f"${nl:.2f}T",
                       (f"▼ {nl3:.1f}% 3m" if (nl3 or 0) < 0 else f"▲ +{nl3:.1f}% 3m") if nl3 is not None else "",
                       "down" if (nl3 or 0) < 0 else "up")
    # stablecoin supply (DefiLlama)
    try:
        pa = requests.get("https://stablecoins.llama.fi/stablecoins?includePrices=false",
                          headers=UA, timeout=TIMEOUT).json()["peggedAssets"]
        total = sum((a.get("circulating", {}) or {}).get("peggedUSD", 0) or 0 for a in pa)
        usdc = next(((a.get("circulating", {}) or {}).get("peggedUSD", 0) for a in pa if a.get("symbol") == "USDC"), 0)
        if total:
            lstats[2] = st(f"${total / 1e9:.0f}B", f"USDC ${usdc / 1e9:.0f}B", "up")
    except Exception as e:
        print(f"  ⚠️  DefiLlama: {e}")
    liq["stats"] = lstats
    out["liquidity"] = liq

    # ── growth (claims, NFP MoM, GDPNow; ISM manual) ─────────────────────────
    g = list(prev.get("growth", [st(""), st(""), st(""), st("")]))
    payems = fred_series("PAYEMS")
    if len(payems) >= 2:
        chg = round(payems[-1][1] - payems[-2][1])  # thousands
        g[1] = st(f"{'+' if chg >= 0 else ''}{chg}K", "→ MoM", "neu")
    claims = fred_series("ICSA")
    if len(claims) >= 2:
        c = claims[-1][1]; pc = claims[-2][1]
        sub, d = _arrow(c, pc)
        g[2] = st(f"{c/1000:.0f}K", f"{sub} {'+' if c>=pc else ''}{round((c-pc)/1000)}K" if pc else "", "down" if c > pc else "up")
    gdpnow = _gdpnow()
    if gdpnow is not None:
        g[3] = st(f"{gdpnow:.1f}%", "Atlanta Fed", "up")
    # ISM has no free API → NY Fed Empire State as an honest regional proxy
    es = fred_series("GACDISA066MSFRBNY")
    if len(es) >= 2:
        v, pv = es[-1][1], es[-2][1]
        sub, _ = _arrow(v, pv)
        g[0] = st(f"{v:+.1f}", "Empire State · NY Fed", "up" if v > 0 else "down")
    out["growth"] = g

    # ── rates (curve + 2Y/10Y/2s10s/HY) ──────────────────────────────────────
    rates = dict(prev.get("rates", {}))
    curve_map = [("1M", "DGS1MO"), ("3M", "DGS3MO"), ("6M", "DGS6MO"), ("2Y", "DGS2"),
                 ("5Y", "DGS5"), ("10Y", "DGS10"), ("30Y", "DGS30")]
    curve = []
    for tenor, sid in curve_map:
        y = fred_last(sid)
        if y is not None:
            curve.append({"tenor": tenor, "y": y})
    if len(curve) >= 5:
        rates["curve"] = curve
    y2, y10 = fred_last("DGS2"), fred_last("DGS10")
    spread, hy = fred_last("T10Y2Y"), fred_last("BAMLH0A0HYM2")
    rstats = list(rates.get("stats", [st(""), st(""), st(""), st("")]))
    if y2 is not None:
        rstats[0] = st(f"{y2:.2f}", rstats[0].get("sub", ""), "neu")
    if y10 is not None:
        rstats[1] = st(f"{y10:.2f}", rstats[1].get("sub", ""), "neu")
    if spread is not None:
        rstats[2] = st(f"{'+' if spread >= 0 else ''}{round(spread*100)}bp",
                       "steepening" if spread > 0 else "inverted", "up" if spread > 0 else "down")
    if hy is not None:
        rstats[3] = st(f"{round(hy*100)}bp", "→ calm" if hy < 4 else "▲ widening", "neu" if hy < 4 else "down")
    rates["stats"] = rstats
    out["rates"] = rates

    # ── regime tape (top strip) — derive all but cut-odds from live data ──────
    mi = load_json("market.json", {}).get("instruments", {})
    prev_tape = {c.get("k"): c for c in prev.get("regime_tape", [])}
    nf_v, nf_d = _next_fomc()
    infl = out.get("inflation", [])
    core = infl[1] if len(infl) > 1 else {}
    tape = []
    if lo is not None and hi is not None:
        tape.append({"k": "Fed Funds", "v": f"{lo:.2f}–{hi:.2f}", "d": "held", "dir": "neu"})
    tape.append({"k": "Next FOMC", "v": nf_v, "d": nf_d, "dir": "neu"})
    if cut_odds:
        f0 = cut_odds[0]
        tape.append({"k": f"Cut odds ({f0['m'].split()[0]})", "v": f"{f0['p']}%",
                     "d": "implied", "dir": "up" if f0["p"] >= 50 else "neu"})
    elif prev_tape.get("Cut odds (Jul)"):
        tape.append(prev_tape["Cut odds (Jul)"])
    if core.get("v"):
        tape.append({"k": "Core CPI", "v": core["v"], "d": "core", "dir": core.get("dir", "neu")})
    if y10 is not None:
        tape.append({"k": "US 10Y", "v": f"{y10:.2f}", "d": "", "dir": "neu"})
    if spread is not None:
        tape.append({"k": "2s10s", "v": f"{'+' if spread >= 0 else ''}{round(spread*100)}",
                     "d": "steepening" if spread > 0 else "inverted", "dir": "up" if spread > 0 else "down"})
    if mi.get("dxy", {}).get("px"):
        tape.append({"k": "DXY", "v": str(mi["dxy"]["px"]), "d": "", "dir": "neu"})
    if mi.get("fg", {}).get("px"):
        fg = mi["fg"]
        tape.append({"k": "F&amp;G", "v": str(fg["px"]), "d": fg.get("chg", ""),
                     "dir": {"up": "up", "dn": "down"}.get(fg.get("dir"), "neu")})
    if len(tape) >= 6:
        out["regime_tape"] = tape

    return out


def _num(s):
    if s is None:
        return None
    try:
        return float(str(s).replace("%", "").replace("$", "").replace("+", "").split("–")[0].split("bp")[0])
    except Exception:
        return None


def _gdpnow():
    """Latest GDPNow forecast from the Atlanta Fed xlsx (needs openpyxl)."""
    try:
        import openpyxl
        r = requests.get("https://www.atlantafed.org/-/media/documents/cqer/researchcq/"
                         "gdpnow/GDPTrackingModelDataAndForecasts.xlsx",
                         headers=UA, timeout=TIMEOUT)
        wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True, read_only=True)
        ws = wb["TrackingHistory"] if "TrackingHistory" in wb.sheetnames else wb.active
        last = None
        for row in ws.iter_rows(values_only=True):
            for c in row:
                if isinstance(c, (int, float)) and -20 < c < 20:
                    last = c
        return round(last, 1) if last is not None else None
    except Exception as e:
        print(f"  ⚠️  GDPNow: {e}")
        return None


if __name__ == "__main__":
    print("Fetching macro regime snapshot…")
    data = build_macro2()
    write_json("macro2.json", data)
    print(f"• macro2: curve {len(data.get('rates', {}).get('curve', []))} pts, "
          f"net-liq series {len(data.get('liquidity', {}).get('net_liq_series', []))}")
